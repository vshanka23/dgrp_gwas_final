#!/usr/bin/env python

import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import argparse
import os
import time

def main(input_gwas, input_vep, input_gene, output_excel):
    # Step 1: Load the data
    gwas = pd.read_csv(input_gwas, sep="\t")
    raw = pd.read_csv(input_vep, sep="\t")
    gene_symbol = pd.read_csv(input_gene, sep="\t")

    #Bug fix for MT variants
    gwas = gwas[~gwas.rs.str.contains("MT_")]

    # Step 2: Filter and merge data
    filt = raw[raw["#Uploaded_variation"].isin(gwas["rs"].tolist())]
    annot = pd.merge(filt, gwas, how='outer', left_on="#Uploaded_variation", right_on="rs")
    
    # Rename gene symbol column for consistency
    gene_symbol = gene_symbol.rename(columns={"primary_FBid": "Gene"})
    
    # Merge with gene symbol data
    annot2 = pd.merge(annot, gene_symbol[['Gene', 'current_symbol']], on="Gene", how='left')
    
    # Step 3: Rename columns
    annot2 = annot2.rename(columns={
        "#Uploaded_variation": "Variant ID",
        "allele1": "Minor Allele",
        "allele0": "Major Allele",
        "n_miss": "Number Missing Lines",
        "af": "Minor Allele Frequency",
        "beta": "Additive Effect (β)",
        "se": "β Standard Error",
        "p_wald": "Wald Test P-Value",
        "current_symbol": "Gene Symbol",
        "Consequence": " Site Class",
        "Extra": "Impact and Distance From Neighboring Genes",
        "Gene": "FlyBase ID"
    })
    
    # Filter out rows containing "FBti" in "FlyBase ID"
    annot2 = annot2[~annot2["FlyBase ID"].str.contains("FBti")]
    
    # Select the final columns
    annot3 = annot2[[
        "Variant ID", "Number Missing Lines", "Minor Allele", "Major Allele", 
        "Minor Allele Frequency", "Additive Effect (β)", "β Standard Error", 
        "Wald Test P-Value", "FlyBase ID", "Gene Symbol", " Site Class", 
        "Impact and Distance From Neighboring Genes"
    ]]
    
    # Replace NaN values with "-"
    annot3 = annot3.replace(np.NaN, '-')

    # Step 4: Round the necessary columns
    annot3["Minor Allele Frequency"] = annot3["Minor Allele Frequency"].round(3)
    annot3["Additive Effect (β)"] = annot3["Additive Effect (β)"].round(3)
    annot3["β Standard Error"] = annot3["β Standard Error"].round(3)

    # Save to Excel
    temp_file = "temp_test.xlsx"
    annot3.to_excel(temp_file, index=False)

    # Delay to ensure file is available on disk
    time.sleep(30)  # Sleep for 30 seconds


    # Step 5: Load the workbook for formatting
    workbook = load_workbook(temp_file)
    sheet = workbook.active

    # Define black line border
    thin_border = Border(
        left=Side(style='thin', color="000000"),
        right=Side(style='thin', color="000000"),
        top=Side(style='thin', color="000000"),
        bottom=Side(style='thin', color="000000")
    )

    # Step 6: Format the header row
    header_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, name="Arial")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    # Step 7: Alternating row colors and italicize "Gene Symbol"
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    no_fill = PatternFill(fill_type=None)

    prev_entry = None
    is_gray = False
    data_alignment = Alignment(horizontal="center", vertical="center")

    for row in sheet.iter_rows(min_row=2):
        cell_value = row[0].value
        
        if cell_value != prev_entry:
            is_gray = not is_gray
            prev_entry = cell_value
        
        fill = gray_fill if is_gray else no_fill
        for i, cell in enumerate(row):
            cell.fill = fill
            cell.alignment = data_alignment
            cell.font = Font(name="Arial", italic=True if i == 9 else False)
            if i == 7:  # Wald Test P-Value in scientific notation
                cell.number_format = '0.00E+00'
            cell.border = thin_border

    # Step 8: Adjust column widths
    column_widths = {
        1: 31, 2: 16, 3: 16, 4: 16, 5: 16, 6: 16, 7: 16,
        8: 16, 9: 21, 10: 21, 11: 51, 12: 183
    }

    for col_num, width in column_widths.items():
        sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = width

    # Save the final formatted file
    workbook.save(output_excel)

    # Remove the intermediary file
    os.remove(temp_file)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Annotate GWAS data using VEP and gene symbol data, and format Excel output.")
    parser.add_argument("-i", "--input_gwas", required=True, help="Path to the input GWAS file")
    parser.add_argument("-v", "--input_vep", required=True, help="Path to the raw VEP file")
    parser.add_argument("-g", "--input_gene", required=True, help="Path to the gene symbol file")
    parser.add_argument("-o", "--output_excel", required=True, help="Path to save the formatted Excel file")

    args = parser.parse_args()

    main(args.input_gwas, args.input_vep, args.input_gene, args.output_excel)